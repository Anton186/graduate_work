"""
Модуль манипуляций Эксель-файлов с помощью библиотеки openpyxl
"""
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils.cell import column_index_from_string


# НАДО ПРОВЕРИТЬ И СКОРРЕКТИРОВАТЬ
def merge_cells_in_column_by_value(file_path, sheet_name, column_letter):
    """
    Функция слияния одинаковых клеток столбцов
    :param file_path:
    :param sheet_name:
    :param column_letter:
    :return:
    """
    workbook = load_workbook(file_path)
    # Выбор листа
    sheet = workbook[sheet_name]
    # Получаем максимальное количество строк и столбцов на листе
    max_row = sheet.max_row
    # Ищем и объединяем ячейки с одинаковыми значениями
    current_value = None
    start_row = 1
    for row in range(1, max_row + 1):
        cell = sheet[column_letter + str(row)]
        if cell.value != current_value:
            if start_row != row - 1:
                sheet.merge_cells(start_row=start_row, end_row=row - 1,
                                  start_column=column_index_from_string(column_letter),
                                  end_column=column_index_from_string(column_letter))
            current_value = cell.value
            start_row = row

    workbook.save(file_path)


def copy_column_names_to_excel(file_path, output_file):  # РАБОТАЕТ В РАМКАХ ФУНКЦИОНАЛА
    wb = load_workbook(file_path)
    wb_sheet = wb.active

    output_wb = Workbook()
    output_sheet = output_wb.active
    values_list = []
    for cell in wb_sheet['1']:
        if cell.value:
            values_list.append(cell.value)

    output_sheet.append(values_list)
    output_wb.save(output_file)
    del values_list[:]


def copy_rows_to_excel_by_value(file_path, target_file, search_value=None, min_date=None, max_date=None):
    source_wb = load_workbook(file_path)
    source_ws = source_wb.active

    target_wb = load_workbook(target_file)
    target_ws = target_wb.active

    # Итерация по всем строкам и столбцам первого Excel-файла + форматирование данных
    if is_row_exists(source_ws, 2):
        if search_value and min_date and max_date:
            for row in source_ws.iter_rows(min_row=2):
                values_list = []
                for cell in row:
                    values_list.append(cell.value)

                temp_list = []
                if search_value in values_list and any(isinstance(value, datetime) for value in values_list):
                    date_value = next((value for value in values_list if isinstance(value, datetime)), None)
                    if min_date <= date_value <= max_date:
                        for value in values_list:
                            if isinstance(value, datetime):
                                value = convert_cell_datetime_value_to_str(value)
                                temp_list.append(value)
                            else:
                                temp_list.append(str(value))
                target_ws.append(temp_list)
                del values_list[:]
                del temp_list[:]


        elif search_value and min_date:
            for row in source_ws.iter_rows(min_row=2):
                values_list = []
                for cell in row:
                    values_list.append(cell.value)

                temp_list = []
                if search_value in values_list and any(isinstance(value, datetime) for value in values_list):
                    date_value = next((value for value in values_list if isinstance(value, datetime)), None)
                    if min_date == date_value:
                        for value in values_list:
                            if isinstance(value, datetime):
                                value = convert_cell_datetime_value_to_str(value)
                                temp_list.append(value)
                            else:
                                temp_list.append(str(value))
                target_ws.append(temp_list)
                del values_list[:]
                del temp_list[:]


        elif search_value:
            for row in source_ws.iter_rows(min_row=2):
                values_list = []
                for cell in row:
                    if isinstance(cell.value, datetime):
                        value = convert_cell_datetime_value_to_str(cell.value)
                        values_list.append(value)
                    else:
                        values_list.append(str(cell.value))
                temp_list = []
                if search_value in values_list:
                    for value in values_list:
                        if isinstance(value, datetime):
                            value = convert_cell_datetime_value_to_str(value)
                            temp_list.append(value)
                        else:
                            temp_list.append(str(value))
                if not len(temp_list) == 0:
                    target_ws.append(temp_list)
                del values_list[:]
                del temp_list[:]


        elif min_date and max_date:
            for row in source_ws.iter_rows(min_row=2):
                values_list = []
                for cell in row:
                    values_list.append(cell.value)

                temp_list = []
                if any(isinstance(value, datetime) for value in values_list):
                    date_value = next((value for value in values_list if isinstance(value, datetime)), None)
                    if min_date <= date_value <= max_date:
                        for value in values_list:
                            if isinstance(value, datetime):
                                value = convert_cell_datetime_value_to_str(value)
                                temp_list.append(value)
                            else:
                                temp_list.append(str(value))
                target_ws.append(temp_list)
                del values_list[:]
                del temp_list[:]


        elif min_date:
            for row in source_ws.iter_rows(min_row=2):
                values_list = []
                for cell in row:
                    values_list.append(cell.value)

                temp_list = []
                if any(isinstance(value, datetime) for value in values_list):
                    date_value = next((value for value in values_list if isinstance(value, datetime)), None)
                    if min_date == date_value:
                        for value in values_list:
                            if isinstance(value, datetime):
                                value = convert_cell_datetime_value_to_str(value)
                                temp_list.append(value)
                            else:
                                temp_list.append(str(value))
                target_ws.append(temp_list)
                del values_list[:]
                del temp_list[:]

        target_wb.save(target_file)

    else:
        print('В файле нет значений!')


def convert_cell_datetime_value_to_str(cell_value: Cell.value):
    """
    Конвертация значения ячейки типа datetime в строку вида 'дд.мм.гггг'
    :param cell_value: значение ячейки
    :return: строка вида "дд.мм.гггг" или None
    """
    if cell_value:
        if isinstance(cell_value, datetime):
            str_value = cell_value.strftime('%d.%m.%Y')
            return str_value
        else:
            print("Значение не является датой")
            return None

    else:
        print("Ячейка пуста")
        return None


def convert_str_to_datetime(data_string: str):
    if data_string:
        if isinstance(data_string, str):
            data_string = datetime.strptime(data_string, '%d.%m.%Y')
            return data_string
        else:
            print("Значение не является строкой:", type(data_string))
            return None

    else:
        print("Ячейка пуста")
        return None


def find_cell(file_path, search_value):
    """
    Поиск ячейки с заданным значением
    :param file_path: путь к эксель-файлу для поиска
    :param search_value: искомое значение
    :return: Cell.cell или None
    """
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook.active

    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value) == search_value:
                cell_coordinate = cell
                return cell_coordinate

    print('Ячейка с заданным значением не найдена')
    return None  # Если ячейка с заданным значением не найдена


def get_search_date():
    start_date = input('Введите дату начала поиска в формате дд.мм.гггг (или оставьте пустым): ')
    if start_date:
        start_date = datetime.strptime(start_date, '%d.%m.%Y')

    final_date = input('Введите дату конца поиска в формате дд.мм.гггг (или оставьте пустым): ')
    if final_date:
        final_date = datetime.strptime(final_date, '%d.%m.%Y')

    if start_date and final_date:
        if start_date == final_date:
            return start_date
        else:
            return start_date, final_date

    if start_date:
        return start_date
    if final_date:
        return final_date

    return None


def get_filename():
    partial_filename = input("Введите имя или часть имени Excel-файла для обработки: ")
    current_directory = os.path.dirname(os.path.abspath(__file__))
    for file_name in os.listdir(current_directory):
        if partial_filename in file_name:
            full_filename = os.path.join(current_directory, file_name)
            print(f'Наименование файла: {file_name}')
            print(f"Путь файла: {full_filename}")
            return file_name
        else:
            print('Совпадение не найдено!')

    print('Файл не найден!')
    return None


def is_row_exists(worksheet, row_index):
    try:
        row = worksheet[row_index]  # Попытка получить строку
        return True  # Если строка существует
    except IndexError:
        return False  # Если строки не существует


def sum_quantity(filename, output_summary):
    # Открываем файл Excel
    wb = load_workbook(filename)
    sheet = wb.active

    wb_out = load_workbook(output_summary)
    ws_out = wb_out.active

    max_rows = sheet.max_row  # Максимальное количество строк в таблице
    max_columns = sheet.max_column  # Максимальное количество столбцов в таблице

    if is_row_exists(sheet, 2):
        if not is_row_exists(sheet, 3):
            print('Нет данных для обработки!')
            return None
    elif not is_row_exists(sheet, 2):
        print('Нет данных для обработки!')
        return None

    values_list = []
    for row_id in range(2, max_rows + 1):
        temp_list = []
        for col_id in range(1, max_columns + 1):
            cell1 = sheet.cell(row=row_id, column=col_id)
            temp_list.append(cell1.value)

        print(temp_list)
        if len(values_list) != 0:
            for lst in values_list:
                flag = True
                for value in range(0, len(lst)):
                    if value == 3:
                        continue
                    if temp_list[value] != lst[value]:
                        flag = False
                        break
                if flag:
                    curr_ind = int(temp_list[3]) + int(lst[3])
                    lst[3] = curr_ind
                else:
                    values_list.append(temp_list)
                    print(values_list)
                    break

        else:
            values_list.append(temp_list)

    wb.close()
    # Закрываем файл Excel

    for lst in values_list:
        ws_out.append(lst)
    wb_out.save(output_summary)
